# -*- coding: utf-8 -*-
# @Time : 2021/11/21 22:10
# @Author : huix
# @File : 01Excel.py


# YAML
# JSON
# EXCEL
# 官网  https://openpyxl.readthedocs.io/en/latest/

# 写
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

ws4 = wb.create_sheet(title = "mysheet")
for i in range(1,31):
    ws4.cell(column=1,row=i).value="test"

wb.save(filename = "test.xlsx")


# 读
from openpyxl import load_workbook
wb1 = load_workbook(filename = "test.xlsx")
sheet_ranges = wb["mysheet"]
print(sheet_ranges['B1'].value)
for i in range(1,31):
    print(sheet_ranges.cell(column=1, row=i).value)